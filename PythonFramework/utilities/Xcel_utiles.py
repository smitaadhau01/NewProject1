import openpyxl

def get_rowCount(file,sheet_name):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.max_row

def readData(file,sheet_name,row_no,col_no):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    return sheet.cell(row=row_no,column=col_no).value

def writeData(file,sheet_name, row_num, col_num , data):
    book = openpyxl.load_workbook(file)
    sheet = book[sheet_name]
    sheet.cell(row=row_num,column=col_num).value = data
    book.save(file)